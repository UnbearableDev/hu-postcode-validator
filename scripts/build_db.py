"""Build hu_postcodes.sqlite from two upstream sources.

DEV-TIME script. Run once before `apify push`. The resulting SQLite gets baked
into the Actor image. Re-run quarterly when Posta updates the catalog.

Sources:
- Magyar Posta official XLSX (postcode → settlement)
- GitHub `tamas-ferenci/IrszHnk` CSV (settlement → county via KSH)

Output:
- data/hu_postcodes.sqlite — schema:
    postcodes(postcode, settlement, settlement_part, settlement_normalized,
              county, jaras_neve, ksh_code)
    bp_districts(postcode, district)
"""

from __future__ import annotations

import sqlite3
import sys
import unicodedata
from pathlib import Path

import openpyxl

POSTA_XLSX = Path("/tmp/hu_postcodes_source.xlsx")
IRSZHNK_CSV = Path("/tmp/irszhnk.csv")
OUT_SQLITE = Path("data/hu_postcodes.sqlite")


def normalize(s: str | None) -> str:
    """NFKD-strip diacritics + lowercase for fuzzy lookup."""
    if not s:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", s).lower() if not unicodedata.combining(c)
    ).strip()


def load_posta(path: Path) -> list[tuple]:
    """Return list of (postcode, settlement, settlement_part) from Posta XLSX."""
    print(f"Loading Posta XLSX: {path}", file=sys.stderr)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Települések"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # empty header
        if i == 1:
            # header row: ('IRSZ', 'Település', 'Településrész')
            continue
        if not row or row[0] is None:
            continue
        postcode = int(row[0])
        settlement = (row[1] or "").strip() if isinstance(row[1], str) else ""
        settlement_part = (row[2] or "").strip() if isinstance(row[2], str) and row[2] else None
        if not settlement:
            continue
        rows.append((postcode, settlement, settlement_part))
    print(f"  Posta: {len(rows)} rows", file=sys.stderr)
    return rows


def load_irszhnk(path: Path) -> dict[str, dict]:
    """Return {normalized_settlement_name: {county, jaras, ksh_code}}.

    IrszHnk CSV is semicolon-delimited, UTF-8, columns include
    Helység.megnevezése, Vármegye.megnevezése, Járás.neve, Helység.KSH.kódja.
    """
    print(f"Loading IrszHnk CSV: {path}", file=sys.stderr)
    if not path.exists():
        print(f"  WARN: {path} does not exist; county data will be NULL", file=sys.stderr)
        return {}

    import csv

    out: dict[str, dict] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        # Print headers once
        if not reader.fieldnames:
            print("  WARN: empty IrszHnk CSV", file=sys.stderr)
            return out
        # Find the columns we care about by name (defensive — they're stable but be explicit)
        col_settlement = next(
            (c for c in reader.fieldnames if c.startswith("Helység.megnevezése")), None
        )
        col_county = next(
            (c for c in reader.fieldnames if c.startswith("Vármegye.megnevezése")), None
        )
        col_jaras = next((c for c in reader.fieldnames if c.startswith("Járás.neve")), None)
        col_ksh = next(
            (c for c in reader.fieldnames if c.startswith("Helység.KSH.kódja")), None
        )
        if not col_settlement:
            print(
                f"  WARN: settlement column not found; have: {reader.fieldnames[:5]}",
                file=sys.stderr,
            )
            return out

        for row in reader:
            settlement = row.get(col_settlement, "").strip()
            if not settlement:
                continue
            key = normalize(settlement)
            out[key] = {
                "county": (row.get(col_county) or "").strip() or None,
                "jaras_neve": (row.get(col_jaras) or "").strip() or None,
                "ksh_code": (row.get(col_ksh) or "").strip() or None,
            }
    print(f"  IrszHnk: {len(out)} unique settlements", file=sys.stderr)
    return out


def load_bp_districts(path: Path) -> dict[int, str]:
    """Return {postcode: district_roman} from the Bp.u. sheet, deduped."""
    print("Loading Budapest district data from Bp.u. sheet", file=sys.stderr)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Bp.u."]
    out: dict[int, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row or row[0] is None:
            continue
        try:
            postcode = int(row[0])
        except (TypeError, ValueError):
            continue
        # KER column is at index 8 (0-based) per header inspection
        district = row[8] if len(row) > 8 and row[8] else None
        if district and postcode not in out:
            out[postcode] = str(district).strip()
    print(f"  Budapest: {len(out)} postcodes with district", file=sys.stderr)
    return out


# Cities that have internal sub-postcodes encoded ONLY in their street sheets
# (the Települések sheet only lists the city's main postcode for these).
# Without this, lookup_postcode(6720..6728) returns not-found for Szeged etc.
_CITY_STREET_SHEETS = {
    "Szeged": "Szeged u.",
    "Miskolc": "Miskolc u.",
    "Debrecen": "Debrecen u.",
    "Pécs": "Pécs u.",
    "Győr": "Győr u.",
}


def load_city_subpostcodes(path: Path) -> dict[str, set[int]]:
    """Return {city_name: {postcode, ...}} from each <City> u. street sheet.

    Each sheet has columns IRSZ, CíMHELY NEVE, JELLEGE, VÁROSRÉSZ, 1.SZÁM,
    1.JEL, 2.SZÁM, 2.JEL. We only need the unique IRSZ values per sheet.
    """
    print("Loading city sub-postcodes from street sheets", file=sys.stderr)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, set[int]] = {}
    for city, sheet_name in _CITY_STREET_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet {sheet_name!r} missing", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        pcs: set[int] = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if not row or row[0] is None:
                continue
            try:
                pcs.add(int(row[0]))
            except (TypeError, ValueError):
                continue
        out[city] = pcs
        print(f"  {city}: {len(pcs)} unique sub-postcodes from {sheet_name}", file=sys.stderr)
    return out


def build(out_path: Path) -> None:
    posta = load_posta(POSTA_XLSX)
    irszhnk = load_irszhnk(IRSZHNK_CSV)
    bp = load_bp_districts(POSTA_XLSX)
    city_subpcs = load_city_subpostcodes(POSTA_XLSX)

    # CRITICAL: Posta's Települések sheet excludes Budapest. Derive Budapest
    # entries from the Bp.u. street sheet (deduped per postcode).
    # Each Budapest postcode becomes a row: (pc, "Budapest", "<district> kerület").
    bp_settlements = [
        (pc, "Budapest", f"{district.rstrip('.')}. kerület")
        for pc, district in bp.items()
    ]
    print(f"  Synthesizing {len(bp_settlements)} Budapest entries for postcodes table", file=sys.stderr)
    posta.extend(bp_settlements)

    # CRITICAL: Posta's Települések lists the MAIN postcode only for Szeged,
    # Miskolc, Debrecen, Pécs, Győr — their internal sub-postcodes
    # (e.g. 6720-6728 for Szeged) live ONLY in the <City> u. street sheets.
    # Synthesize one row per (city, sub-postcode) that is not already present
    # in posta, so lookup_postcode and validate_address resolve them.
    existing_pcs_per_city: dict[str, set[int]] = {}
    for pc, city, _part in posta:
        existing_pcs_per_city.setdefault(city, set()).add(pc)
    synthesized_subpc = 0
    for city, pcs in city_subpcs.items():
        already = existing_pcs_per_city.get(city, set())
        for pc in sorted(pcs):
            if pc in already:
                continue
            posta.append((pc, city, None))
            synthesized_subpc += 1
    print(
        f"  Synthesizing {synthesized_subpc} city sub-postcode entries (Szeged/Miskolc/Debrecen/Pécs/Győr)",
        file=sys.stderr,
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists():
        out_path.unlink()

    print(f"Writing SQLite: {out_path}", file=sys.stderr)
    conn = sqlite3.connect(out_path)
    try:
        c = conn.cursor()
        c.executescript(
            """
            CREATE TABLE postcodes (
                postcode INTEGER NOT NULL,
                settlement TEXT NOT NULL,
                settlement_part TEXT,
                settlement_normalized TEXT NOT NULL,
                county TEXT,
                jaras_neve TEXT,
                ksh_code TEXT
            );
            CREATE INDEX idx_postcode ON postcodes(postcode);
            CREATE INDEX idx_settlement_norm ON postcodes(settlement_normalized);
            CREATE INDEX idx_county ON postcodes(county);

            CREATE TABLE bp_districts (
                postcode INTEGER PRIMARY KEY,
                district TEXT NOT NULL
            );
            CREATE INDEX idx_bp_district ON bp_districts(district);
            """
        )

        # Insert postcodes — JOIN to IrszHnk by normalized settlement name.
        # For Budapest entries (synthesized above), county is hardcoded since
        # IrszHnk treats Budapest as a single Vármegye-level entity.
        joined = 0
        for postcode, settlement, part in posta:
            norm = normalize(settlement)
            enriched = irszhnk.get(norm) or {}
            if enriched:
                joined += 1
            # Hardcode county for Budapest postcodes that don't match IrszHnk
            if settlement == "Budapest" and not enriched.get("county"):
                county = "Budapest"
            else:
                county = enriched.get("county")
            c.execute(
                """INSERT INTO postcodes
                   (postcode, settlement, settlement_part, settlement_normalized,
                    county, jaras_neve, ksh_code)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    postcode,
                    settlement,
                    part,
                    norm,
                    county,
                    enriched.get("jaras_neve"),
                    enriched.get("ksh_code"),
                ),
            )
        print(f"  postcodes inserted: {len(posta)}, county-enriched: {joined}", file=sys.stderr)

        # Insert Budapest districts
        for postcode, district in bp.items():
            c.execute(
                "INSERT OR REPLACE INTO bp_districts (postcode, district) VALUES (?, ?)",
                (postcode, district),
            )
        print(f"  bp_districts inserted: {len(bp)}", file=sys.stderr)

        conn.commit()

        # Sanity counts
        n_pc = c.execute("SELECT COUNT(*) FROM postcodes").fetchone()[0]
        n_county = c.execute(
            "SELECT COUNT(*) FROM postcodes WHERE county IS NOT NULL"
        ).fetchone()[0]
        n_bp = c.execute("SELECT COUNT(*) FROM bp_districts").fetchone()[0]
        size_mb = out_path.stat().st_size / 1_000_000
        print(
            f"\nResult: {n_pc} postcodes ({n_county} with county), "
            f"{n_bp} BP-district mappings, {size_mb:.2f} MB",
            file=sys.stderr,
        )
    finally:
        conn.close()


if __name__ == "__main__":
    build(OUT_SQLITE)
